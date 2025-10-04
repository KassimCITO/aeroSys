from flask import Blueprint, render_template, request, make_response, jsonify
from flask_login import login_required
from models import db, Aeronave, Piloto, Vuelo, Confirmacion, Usuario
from datetime import datetime, date
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import os

reports_bp = Blueprint('reports', __name__)

def generate_pdf_report(title, data, headers, filename):
    """Genera un reporte PDF con los datos proporcionados"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    # Contenido
    story = []
    
    # Título
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 12))
    
    # Fecha de generación
    story.append(Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    if data:
        # Crear tabla
        table_data = [headers] + data
        table = Table(table_data)
        
        # Estilo de la tabla
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
    else:
        story.append(Paragraph("No hay datos para mostrar", styles['Normal']))
    
    # Construir PDF
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_excel_report(title, data, headers, filename):
    """Genera un reporte Excel con los datos proporcionados"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    # Estilos
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    data_font = Font(name='Arial', size=10)
    
    title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center_alignment
    ws.row_dimensions[1].height = 30
    
    # Fecha de generación
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws['A2'].font = data_font
    ws['A2'].alignment = center_alignment
    ws.row_dimensions[2].height = 20
    
    # Espacio
    ws.row_dimensions[3].height = 10
    
    if data:
        # Headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Datos
        for row_idx, row_data in enumerate(data, 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = center_alignment
                cell.border = thin_border
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(4, len(data) + 5):
                try:
                    if ws[f'{column_letter}{row}'].value:
                        max_length = max(max_length, len(str(ws[f'{column_letter}{row}'].value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    else:
        ws['A4'] = "No hay datos para mostrar"
        ws['A4'].font = data_font
        ws['A4'].alignment = center_alignment
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@reports_bp.route('/aeronaves')
@login_required
def report_aeronaves():
    """Genera reporte de aeronaves en PDF o Excel"""
    # Obtener filtros
    fabricante = request.args.get('fabricante', '')
    tipo = request.args.get('tipo_aeronave', '')
    formato = request.args.get('formato', 'pdf')  # pdf o excel
    
    # Consulta con filtros
    query = Aeronave.query
    if fabricante:
        query = query.filter(Aeronave.fabricante.ilike(f'%{fabricante}%'))
    if tipo:
        query = query.filter(Aeronave.tipo_aeronave == tipo)
    
    aeronaves = query.all()
    
    # Preparar datos
    data = []
    for a in aeronaves:
        # Para el reporte, mostramos "Con imagen" o "Sin imagen" en lugar del nombre del archivo
        imagen_status = "Con imagen" if a.imagen else "Sin imagen"
        data.append([
            a.matricula,
            a.modelo,
            a.fabricante,
            str(a.capacidad) if a.capacidad else 'N/A',
            a.tipo_aeronave or 'N/A',
            imagen_status
        ])
    
    headers = ['Matrícula', 'Modelo', 'Fabricante', 'Capacidad', 'Tipo', 'Estado Imagen']
    title = f"Reporte de Aeronaves - {len(data)} registros"
    
    if fabricante:
        title += f" (Fabricante: {fabricante})"
    if tipo:
        title += f" (Tipo: {tipo})"
    
    # Generar reporte según formato
    if formato.lower() == 'excel':
        buffer = generate_excel_report(title, data, headers, 'aeronaves.xlsx')
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=aeronaves_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    else:  # PDF por defecto
        buffer = generate_pdf_report(title, data, headers, 'aeronaves.pdf')
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=aeronaves_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response

@reports_bp.route('/pilotos')
@login_required
def report_pilotos():
    """Genera reporte de pilotos en PDF o Excel"""
    # Obtener filtros
    tipo_licencia = request.args.get('tipo_licencia', '')
    nacionalidad = request.args.get('nacionalidad', '')
    formato = request.args.get('formato', 'pdf')  # pdf o excel
    
    # Consulta con filtros
    query = Piloto.query
    # Nota: El modelo Piloto no tiene campos tipo_licencia ni nacionalidad
    
    pilotos = query.all()
    
    # Preparar datos
    data = []
    for p in pilotos:
        data.append([
            p.nombre,
            p.licencia,
            'N/A',  # El modelo Piloto no tiene campo tipo_licencia
            str(p.horas_vuelo) if p.horas_vuelo else '0',
            'N/A'   # El modelo Piloto no tiene campo nacionalidad
        ])
    
    headers = ['Nombre', 'Licencia', 'Tipo Licencia', 'Horas Vuelo', 'Nacionalidad']
    title = f"Reporte de Pilotos - {len(data)} registros"
    
    if tipo_licencia:
        title += f" (Tipo: {tipo_licencia})"
    if nacionalidad:
        title += f" (Nacionalidad: {nacionalidad})"
    
    # Generar reporte según formato
    if formato.lower() == 'excel':
        buffer = generate_excel_report(title, data, headers, 'pilotos.xlsx')
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=pilotos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    else:  # PDF por defecto
        buffer = generate_pdf_report(title, data, headers, 'pilotos.pdf')
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=pilotos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response

@reports_bp.route('/vuelos')
@login_required
def report_vuelos():
    """Genera reporte de vuelos en PDF o Excel"""
    # Obtener filtros
    estado = request.args.get('estado', '')
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    formato = request.args.get('formato', 'pdf')  # pdf o excel
    
    # Consulta con filtros
    query = Vuelo.query
    # Nota: El modelo Vuelo no tiene campo 'estado', se filtra por confirmaciones si es necesario
    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            query = query.filter(Vuelo.fecha_salida >= fecha_desde_dt)
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            query = query.filter(Vuelo.fecha_salida <= fecha_hasta_dt)
        except ValueError:
            pass
    
    vuelos = query.all()
    
    # Preparar datos
    data = []
    for v in vuelos:
        # Obtener aeronave y piloto usando las claves foráneas correctas
        aeronave = v.aeronave
        piloto = v.piloto
        
        data.append([
            v.numero_vuelo,
            v.origen,
            v.destino,
            v.fecha_salida.strftime('%d/%m/%Y %H:%M') if v.fecha_salida else 'N/A',
            v.fecha_llegada.strftime('%d/%m/%Y %H:%M') if v.fecha_llegada else 'N/A',
            v.observaciones or 'N/A',
            aeronave.matricula if aeronave else 'N/A',
            piloto.nombre if piloto else 'N/A'
        ])
    
    headers = ['Número Vuelo', 'Origen', 'Destino', 'Salida', 'Llegada', 'Observaciones', 'Aeronave', 'Piloto']
    title = f"Reporte de Vuelos - {len(data)} registros"
    
    if estado:
        title += f" (Estado: {estado})"
    if fecha_desde or fecha_hasta:
        title += f" (Período: {fecha_desde or 'Inicio'} - {fecha_hasta or 'Actual'})"
    
    # Generar reporte según formato
    if formato.lower() == 'excel':
        buffer = generate_excel_report(title, data, headers, 'vuelos.xlsx')
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=vuelos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    else:  # PDF por defecto
        buffer = generate_pdf_report(title, data, headers, 'vuelos.pdf')
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=vuelos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response

@reports_bp.route('/confirmaciones')
@login_required
def report_confirmaciones():
    """Genera reporte de confirmaciones en PDF o Excel"""
    # Obtener filtros
    estado = request.args.get('estado', '')
    vuelo_id = request.args.get('vuelo_id', '')
    formato = request.args.get('formato', 'pdf')  # pdf o excel
    
    # Consulta con filtros
    query = Confirmacion.query
    if estado:
        query = query.filter(Confirmacion.estado == estado)
    if vuelo_id:
        query = query.filter(Confirmacion.vuelo_id == int(vuelo_id))
    
    confirmaciones = query.all()
    
    # Preparar datos
    data = []
    for c in confirmaciones:
        # Obtener vuelo usando la relación
        vuelo = c.vuelo
        
        data.append([
            'N/A',  # El modelo Confirmacion no tiene campo nombre_pasajero
            'N/A',  # El modelo Confirmacion no tiene campo documento_identidad
            'N/A',  # El modelo Confirmacion no tiene campo asiento
            c.estado,
            vuelo.numero_vuelo if vuelo else f'Vuelo {c.vuelo_id}',
            vuelo.origen + ' - ' + vuelo.destino if vuelo else 'N/A'
        ])
    
    headers = ['Pasajero', 'Documento', 'Asiento', 'Estado', 'Número Vuelo', 'Ruta']
    title = f"Reporte de Confirmaciones - {len(data)} registros"
    
    if estado:
        title += f" (Estado: {estado})"
    if vuelo_id:
        title += f" (Vuelo: {vuelo_id})"
    
    # Generar reporte según formato
    if formato.lower() == 'excel':
        buffer = generate_excel_report(title, data, headers, 'confirmaciones.xlsx')
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=confirmaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    else:  # PDF por defecto
        buffer = generate_pdf_report(title, data, headers, 'confirmaciones.pdf')
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=confirmaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response

@reports_bp.route('/dashboard')
@login_required
def report_dashboard():
    """Genera reporte del dashboard en PDF o Excel con estadísticas generales"""
    formato = request.args.get('formato', 'pdf')  # pdf o excel
    
    # Obtener estadísticas
    total_aeronaves = Aeronave.query.count()
    total_pilotos = Piloto.query.count()
    total_vuelos = Vuelo.query.count()
    total_confirmaciones = Confirmacion.query.count()
    
    # Vuelos por estado (usando confirmaciones)
    vuelos_por_estado = db.session.query(Confirmacion.estado, db.func.count(Confirmacion.confirmacion_id)).group_by(Confirmacion.estado).all()
    
    # Preparar datos
    data = [
        ['Total de Aeronaves', str(total_aeronaves)],
        ['Total de Pilotos', str(total_pilotos)],
        ['Total de Vuelos', str(total_vuelos)],
        ['Total de Confirmaciones', str(total_confirmaciones)],
        ['', ''],
        ['Confirmaciones por Estado:', ''],
    ]
    
    for estado, count in vuelos_por_estado:
        data.append([estado, str(count)])
    
    headers = ['Métrica', 'Valor']
    title = "Reporte de Estadísticas Generales"
    
    # Generar reporte según formato
    if formato.lower() == 'excel':
        buffer = generate_excel_report(title, data, headers, 'dashboard.xlsx')
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=dashboard_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    else:  # PDF por defecto
        buffer = generate_pdf_report(title, data, headers, 'dashboard.pdf')
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=dashboard_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response

@reports_bp.route('/')
@login_required
def reports_index():
    """Página principal de reportes"""
    return render_template('reports.html')

@reports_bp.route('/filtros/aeronaves')
@login_required
def filtros_aeronaves():
    """Muestra formulario de filtros para reporte de aeronaves"""
    # Obtener opciones únicas para los filtros
    fabricantes = db.session.query(Aeronave.fabricante).distinct().filter(Aeronave.fabricante.isnot(None)).all()
    tipos = db.session.query(Aeronave.tipo_aeronave).distinct().filter(Aeronave.tipo_aeronave.isnot(None)).all()
    
    return render_template('reports/filtros_aeronaves.html', 
                         fabricantes=[f[0] for f in fabricantes],
                         tipos=[t[0] for t in tipos])

@reports_bp.route('/filtros/pilotos')
@login_required
def filtros_pilotos():
    """Muestra formulario de filtros para reporte de pilotos"""
    # Obtener opciones únicas para los filtros
    # Nota: El modelo Piloto no tiene campos tipo_licencia ni nacionalidad
    tipos_licencia = []
    nacionalidades = []
    
    return render_template('reports/filtros_pilotos.html',
                         tipos_licencia=tipos_licencia,
                         nacionalidades=nacionalidades)

@reports_bp.route('/filtros/vuelos')
@login_required
def filtros_vuelos():
    """Muestra formulario de filtros para reporte de vuelos"""
    # Obtener opciones únicas para los filtros
    # Nota: El modelo Vuelo no tiene campo estado
    estados = []
    
    return render_template('reports/filtros_vuelos.html',
                         estados=estados)

@reports_bp.route('/filtros/confirmaciones')
@login_required
def filtros_confirmaciones():
    """Muestra formulario de filtros para reporte de confirmaciones"""
    # Obtener opciones únicas para los filtros
    estados = db.session.query(Confirmacion.estado).distinct().filter(Confirmacion.estado.isnot(None)).all()
    vuelos = db.session.query(Vuelo.vuelo_id, Vuelo.numero_vuelo).all()
    
    return render_template('reports/filtros_confirmaciones.html',
                         estados=[e[0] for e in estados],
                         vuelos=vuelos)


